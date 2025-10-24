# view_celle_multiple.py
import json
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from gestione_aree_frame_async import AsyncRunner

def _json_obj(res):
    if isinstance(res, str):
        try:
            res = json.loads(res)
        except Exception as ex:
            raise RuntimeError(f"Risposta non JSON: {ex}\nRaw: {res!r}")
    if isinstance(res, dict) and "error" in res:
        err = res.get("error") or "Errore sconosciuto"
        detail = res.get("sql") or ""
        raise RuntimeError(f"{err}\n{detail}")
    return res if isinstance(res, dict) else {"rows": res}

UBI_B = (
    "UPPER("
    "  CONCAT("
    "    RTRIM(b.Corsia), '.', RTRIM(CAST(b.Colonna AS varchar(32))), '.', RTRIM(CAST(b.Fila AS varchar(32)))"
    "  )"
    ")"
)

BASE_CTE = """
WITH base AS (
  SELECT
    g.IDCella,
    g.BarcodePallet,
    RTRIM(c.Corsia) AS Corsia,
    c.Colonna,
    c.Fila
  FROM dbo.XMag_GiacenzaPallet AS g
  JOIN dbo.Celle AS c ON c.ID = g.IDCella
  WHERE g.IDCella <> 9999 AND RTRIM(c.Corsia) <> '7G'
)
"""

SQL_CORSIE = BASE_CTE + """
, dup_celle AS (
  SELECT IDCCella = b.IDCella
  FROM base b
  GROUP BY b.IDCella
  HAVING COUNT(DISTINCT b.BarcodePallet) > 1
)
SELECT DISTINCT b.Corsia
FROM base b
WHERE EXISTS (SELECT 1 FROM dup_celle d WHERE d.IDCCella = b.IDCella)
ORDER BY b.Corsia;
"""

SQL_CELLE_DUP_PER_CORSIA = BASE_CTE + f"""
, dup_celle AS (
  SELECT b.IDCella, COUNT(DISTINCT b.BarcodePallet) AS NumUDC
  FROM base b
  GROUP BY b.IDCella
  HAVING COUNT(DISTINCT b.BarcodePallet) > 1
)
SELECT dc.IDCella,
       {UBI_B} AS Ubicazione,
       b.Colonna, b.Fila, b.Corsia,
       dc.NumUDC
FROM dup_celle dc
JOIN base b ON b.IDCella = dc.IDCella
WHERE b.Corsia = RTRIM(:corsia)
GROUP BY dc.IDCella, {UBI_B}, b.Colonna, b.Fila, b.Corsia, dc.NumUDC
ORDER BY b.Colonna, b.Fila;
"""

SQL_PALLET_IN_CELLA = BASE_CTE + """
SELECT
  b.BarcodePallet AS Pallet,
  ta.Descrizione,
  ta.Lotto
FROM base b
OUTER APPLY (
  SELECT TOP (1) t.Descrizione, t.Lotto
  FROM dbo.vXTracciaProdotti AS t
  WHERE t.Pallet = b.BarcodePallet COLLATE Latin1_General_CI_AS
  ORDER BY t.Lotto
) AS ta
WHERE b.IDCella = :idcella
GROUP BY b.BarcodePallet, ta.Descrizione, ta.Lotto
ORDER BY b.BarcodePallet;
"""

SQL_RIEPILOGO_PERCENTUALI = BASE_CTE + """
, tot AS (
  SELECT b.Corsia, COUNT(DISTINCT b.IDCella) AS TotCelle
  FROM base b GROUP BY b.Corsia
),
dup_celle AS (
  SELECT b.Corsia, b.IDCella
  FROM base b
  GROUP BY b.Corsia, b.IDCella
  HAVING COUNT(DISTINCT b.BarcodePallet) > 1
),
per_corsia AS (
  SELECT t.Corsia, t.TotCelle, COALESCE(d.CelleMultiple, 0) AS CelleMultiple
  FROM tot t
  LEFT JOIN (
     SELECT Corsia, COUNT(IDCella) AS CelleMultiple
     FROM dup_celle GROUP BY Corsia
  ) d ON d.Corsia = t.Corsia
),
unione AS (
  SELECT Corsia, TotCelle, CelleMultiple,
         CAST(100.0 * CelleMultiple / NULLIF(TotCelle, 0) AS decimal(5,2)) AS Percentuale,
         CAST(0 AS int) AS Ord
  FROM per_corsia
  UNION ALL
  SELECT 'TOTALE' AS Corsia,
         SUM(TotCelle), SUM(CelleMultiple),
         CAST(100.0 * SUM(CelleMultiple) / NULLIF(SUM(TotCelle), 0) AS decimal(5,2)),
         CAST(1 AS int) AS Ord
  FROM per_corsia
)
SELECT Corsia, TotCelle, CelleMultiple, Percentuale
FROM unione
ORDER BY Ord, Corsia;
"""

class CelleMultipleWindow(tk.Toplevel):
    def __init__(self, root, db_client, runner: AsyncRunner | None = None):
        super().__init__(root)
        self.title("Celle con più pallet")
        self.geometry("1100x700"); self.minsize(900,550); self.resizable(True, True)

        self.db = db_client
        self.runner = runner or AsyncRunner(self)

        self._build_layout()
        self._bind_events()
        self.refresh_all()

    def _build_layout(self):
        self.grid_rowconfigure(0, weight=5)
        self.grid_rowconfigure(1, weight=70)
        self.grid_rowconfigure(2, weight=25, minsize=160)
        self.grid_columnconfigure(0, weight=1)

        toolbar = ttk.Frame(self); toolbar.grid(row=0, column=0, sticky="nsew")
        ttk.Button(toolbar, text="Aggiorna", command=self.refresh_all).pack(side="left", padx=6, pady=4)
        ttk.Button(toolbar, text="Espandi tutto", command=self.expand_all).pack(side="left", padx=6, pady=4)
        ttk.Button(toolbar, text="Comprimi tutto", command=self.collapse_all).pack(side="left", padx=6, pady=4)
        ttk.Button(toolbar, text="Esporta in XLSX", command=self.export_to_xlsx).pack(side="left", padx=6, pady=4)

        f = ttk.Frame(self); f.grid(row=1, column=0, sticky="nsew", padx=6, pady=(0,6))
        f.grid_rowconfigure(0, weight=1); f.grid_columnconfigure(0, weight=1)
        self.tree = ttk.Treeview(f, columns=("col2","col3"), show="tree headings", selectmode="browse")
        self.tree.heading("#0", text="Nodo"); self.tree.heading("col2", text="Descrizione"); self.tree.heading("col3", text="Lotto")
        y = ttk.Scrollbar(f, orient="vertical", command=self.tree.yview)
        x = ttk.Scrollbar(f, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=y.set, xscrollcommand=x.set)
        self.tree.grid(row=0, column=0, sticky="nsew"); y.grid(row=0, column=1, sticky="ns"); x.grid(row=1, column=0, sticky="ew")

        sumf = ttk.LabelFrame(self, text="Riepilogo % celle multiple per corsia")
        sumf.grid(row=2, column=0, sticky="nsew", padx=6, pady=(0,6))
        sumf.grid_rowconfigure(0, weight=1); sumf.grid_columnconfigure(0, weight=1)
        self.sum_tbl = ttk.Treeview(sumf, columns=("Corsia","TotCelle","CelleMultiple","Percentuale"), show="headings")
        for k,t,w,a in (("Corsia","Corsia",100,"center"),
                        ("TotCelle","Totale celle",120,"e"),
                        ("CelleMultiple",">1 UDC",120,"e"),
                        ("Percentuale","%",80,"e")):
            self.sum_tbl.heading(k, text=t); self.sum_tbl.column(k, width=w, anchor=a)
        y2 = ttk.Scrollbar(sumf, orient="vertical", command=self.sum_tbl.yview)
        x2 = ttk.Scrollbar(sumf, orient="horizontal", command=self.sum_tbl.xview)
        self.sum_tbl.configure(yscrollcommand=y2.set, xscrollcommand=x2.set)
        self.sum_tbl.grid(row=0, column=0, sticky="nsew"); y2.grid(row=0, column=1, sticky="ns"); x2.grid(row=1, column=0, sticky="ew")

    def _bind_events(self):
        self.tree.bind("<<TreeviewOpen>>", self._on_open_node)

    def refresh_all(self):
        self._load_corsie(); self._load_riepilogo()

    def _load_corsie(self):
        self.tree.delete(*self.tree.get_children())
        async def _q(db): return await db.query_json(SQL_CORSIE, as_dict_rows=True)
        self.runner.run(_q(self.db), self._fill_corsie, lambda e: messagebox.showerror("Errore", str(e), parent=self))

    def _fill_corsie(self, res):
        rows = _json_obj(res).get("rows", [])
        for r in rows:
            corsia = r.get("Corsia"); 
            if not corsia: continue
            node_id = f"corsia:{corsia}"
            self.tree.insert("", "end", iid=node_id, text=f"Corsia {corsia}", values=("", ""), open=False, tags=("corsia",))
            self.tree.insert(node_id, "end", iid=f"{node_id}::lazy", text="...", values=("", ""))

    def _on_open_node(self, _evt):
        sel = self.tree.focus()
        if not sel: return
        if sel.startswith("corsia:"):
            lazy_id = f"{sel}::lazy"
            if lazy_id in self.tree.get_children(sel):
                self.tree.delete(lazy_id)
                corsia = sel.split(":",1)[1]
                self._load_celle_for_corsia(sel, corsia)
        elif sel.startswith("cella:"):
            lazy_id = f"{sel}::lazy"
            if lazy_id in self.tree.get_children(sel):
                self.tree.delete(lazy_id)
                idcella = int(sel.split(":",1)[1])
                for child in self.tree.get_children(sel):
                    self.tree.delete(child)
                self._load_pallet_for_cella(sel, idcella)

    def _load_celle_for_corsia(self, parent_iid, corsia):
        async def _q(db): return await db.query_json(SQL_CELLE_DUP_PER_CORSIA, params={"corsia": corsia}, as_dict_rows=True)
        self.runner.run(_q(self.db), lambda res: self._fill_celle(parent_iid, res),
                        lambda e: messagebox.showerror("Errore", str(e), parent=self))

    def _fill_celle(self, parent_iid, res):
        rows = _json_obj(res).get("rows", [])
        if not rows:
            self.tree.insert(parent_iid, "end", text="(nessuna cella con >1 UDC)", values=("", "")); return
        for r in rows:
            idc = r["IDCella"]; ubi = r["Ubicazione"]; corsia = r.get("Corsia"); num = r.get("NumUDC", 0)
            node_id = f"cella:{idc}"; label = f"{ubi}  [x{num}]"
            if self.tree.exists(node_id):
                self.tree.item(node_id, text=label, values=(f"IDCella {idc}", ""))
            else:
                self.tree.insert(parent_iid, "end", iid=node_id, text=label,
                                 values=(f"IDCella {idc}", ""), open=False, tags=("cella", f"corsia:{corsia}"))
            if not any(ch.endswith("::lazy") for ch in self.tree.get_children(node_id)):
                self.tree.insert(node_id, "end", iid=f"{node_id}::lazy", text="...", values=("", ""))

    def _load_pallet_for_cella(self, parent_iid, idcella: int):
        async def _q(db): return await db.query_json(SQL_PALLET_IN_CELLA, params={"idcella": idcella}, as_dict_rows=True)
        self.runner.run(_q(self.db), lambda res: self._fill_pallet(parent_iid, res),
                        lambda e: messagebox.showerror("Errore", str(e), parent=self))

    def _fill_pallet(self, parent_iid, res):
        rows = _json_obj(res).get("rows", [])
        if not rows:
            self.tree.insert(parent_iid, "end", text="(nessun pallet)", values=("", "")); return
        parent_tags = self.tree.item(parent_iid, "tags") or ()
        corsia_tag = next((t for t in parent_tags if t.startswith("corsia:")), None)
        corsia_val = corsia_tag.split(":",1)[1] if corsia_tag else ""
        cella_ubi = self.tree.item(parent_iid, "text")
        idcella_txt = self.tree.item(parent_iid, "values")[0]
        idcella_num = int(idcella_txt.split()[-1]) if idcella_txt else None

        for r in rows:
            pallet = r.get("Pallet", ""); desc = r.get("Descrizione", ""); lotto = r.get("Lotto", "")
            leaf_id = f"pallet:{idcella_num}:{pallet}"
            if self.tree.exists(leaf_id):
                self.tree.item(leaf_id, text=str(pallet), values=(desc, lotto)); continue
            self.tree.insert(parent_iid, "end", iid=leaf_id, text=str(pallet),
                             values=(desc, lotto),
                             tags=("pallet", f"corsia:{corsia_val}", f"ubicazione:{cella_ubi}", f"idcella:{idcella_num}"))

    def _load_riepilogo(self):
        async def _q(db): return await db.query_json(SQL_RIEPILOGO_PERCENTUALI, as_dict_rows=True)
        self.runner.run(_q(self.db), self._fill_riepilogo, lambda e: messagebox.showerror("Errore", str(e), parent=self))

    def _fill_riepilogo(self, res):
        rows = _json_obj(res).get("rows", [])
        for i in self.sum_tbl.get_children(): self.sum_tbl.delete(i)
        for r in rows:
            self.sum_tbl.insert("", "end", values=(r.get("Corsia"), r.get("TotCelle",0),
                                                   r.get("CelleMultiple",0), f"{r.get('Percentuale',0):.2f}"))

    def expand_all(self):
        for iid in self.tree.get_children(""):
            self.tree.item(iid, open=True)
            if f"{iid}::lazy" in self.tree.get_children(iid):
                self.tree.delete(f"{iid}::lazy")
                corsia = iid.split(":",1)[1]
                self._load_celle_for_corsia(iid, corsia)

    def collapse_all(self):
        for iid in self.tree.get_children(""):
            self.tree.item(iid, open=False)

    def export_to_xlsx(self):
        ts = datetime.now().strftime("%d_%m_%Y_%H-%M")
        default_name = f"esportazione_celle_udc_multiple_{ts}.xlsx"
        fname = filedialog.asksaveasfilename(parent=self, title="Esporta in Excel",
                                             defaultextension=".xlsx",
                                             filetypes=[("Excel Workbook","*.xlsx")],
                                             initialfile=default_name)
        if not fname: return
        try:
            wb = Workbook()
            ws_det = wb.active; ws_det.title = "Dettaglio"
            ws_sum = wb.create_sheet("Riepilogo")
            det_headers = ["Corsia", "Ubicazione", "IDCella", "Pallet", "Descrizione", "Lotto"]
            sum_headers = ["Corsia", "TotCelle", "CelleMultiple", "Percentuale"]
            def _hdr(ws, headers):
                for j,h in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=j, value=h)
                    cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center", vertical="center")
            _hdr(ws_det, det_headers); _hdr(ws_sum, sum_headers)

            r = 2
            for corsia_node in self.tree.get_children(""):
                for cella_node in self.tree.get_children(corsia_node):
                    for pallet_node in self.tree.get_children(cella_node):
                        tags = self.tree.item(pallet_node, "tags") or ()
                        if "pallet" not in tags: continue
                        corsia = next((t.split(":",1)[1] for t in tags if t.startswith("corsia:")), "")
                        ubi = next((t.split(":",1)[1] for t in tags if t.startswith("ubicazione:")), "")
                        idcella = next((t.split(":",1)[1] for t in tags if t.startswith("idcella:")), "")
                        pallet = self.tree.item(pallet_node, "text")
                        desc, lotto = self.tree.item(pallet_node, "values")
                        for j,v in enumerate([corsia, ubi, idcella, pallet, desc, lotto], start=1):
                            ws_det.cell(row=r, column=j, value=v)
                        r += 1

            r2 = 2
            for iid in self.sum_tbl.get_children(""):
                vals = self.sum_tbl.item(iid, "values")
                for j, v in enumerate(vals, start=1):
                    ws_sum.cell(row=r2, column=j, value=v)
                r2 += 1

            def _autosize(ws):
                widths = {}
                for row in ws.iter_rows(values_only=True):
                    for j, val in enumerate(row, start=1):
                        val_s = "" if val is None else str(val)
                        widths[j] = max(widths.get(j, 0), len(val_s))
                from openpyxl.utils import get_column_letter
                for j, w in widths.items():
                    ws.column_dimensions[get_column_letter(j)].width = min(max(w + 2, 10), 60)

            _autosize(ws_det); _autosize(ws_sum)
            wb.save(fname); messagebox.showinfo("Esportazione completata", f"File creato:\n{fname}", parent=self)
        except Exception as ex:
            messagebox.showerror("Errore esportazione", str(ex), parent=self)

def open_celle_multiple_window(root: tk.Tk, db_client, runner: AsyncRunner | None = None):
    win = CelleMultipleWindow(root, db_client, runner=runner); win.lift(); win.focus_set(); return win
