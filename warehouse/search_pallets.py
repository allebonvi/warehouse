# search_pallets.py
# Finestra di ricerca su UDC / Lotto / Codice Prodotto
# - Tre campi: UDC, Lotto, Codice Prodotto (AND tra i campi valorizzati)
# - Ricerca su TUTTE le celle, incluse IDCella=9999 e corsia 7G
# - Risultati in una griglia: IDCella, Ubicazione, UDC, Lotto, Codice, Descrizione
# - Se la ricerca restituisce >0 righe, i campi input vengono svuotati

from __future__ import annotations
import tkinter as tk
from tkinter import ttk, messagebox

from gestione_aree_frame_async import BusyOverlay, AsyncRunner
from tkinter import filedialog

# opzionale export xlsx
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    _HAS_XLSX = True
except Exception:
    _HAS_XLSX = False

# opzionale: usare tksheet per avere griglie con bordi di cella
try:
    from tksheet import Sheet
except Exception:
    Sheet = None

SQL_SEARCH = r"""
WITH BASE AS (
    SELECT
        g.IDCella,
        -- forza stringa per ricerche LIKE
        CONCAT(g.BarcodePallet, '') AS UDC,
        c.Corsia,
        c.Colonna,
        c.Fila
    FROM dbo.XMag_GiacenzaPallet AS g
    LEFT JOIN dbo.Celle AS c ON c.ID = g.IDCella
    -- NB: qui NON escludiamo IDCella=9999 né '7G'
),
JOINED AS (
    SELECT 
        b.IDCella,
        b.UDC,
        b.Corsia,
        b.Colonna,
        b.Fila,
        t.Lotto,
        t.Prodotto,
        t.Descrizione
    FROM BASE b
    LEFT JOIN dbo.vXTracciaProdotti AS t
      ON t.Pallet COLLATE Latin1_General_CI_AS = LEFT(b.UDC, 6) COLLATE Latin1_General_CI_AS
)
SELECT 
    j.IDCella,
    UPPER(
        CONCAT(
            COALESCE(LTRIM(RTRIM(j.Corsia)), 'NA'), '.',
            COALESCE(LTRIM(RTRIM(CAST(j.Colonna AS varchar(32)))), 'NA'), '.',
            COALESCE(LTRIM(RTRIM(CAST(j.Fila AS varchar(32)))), 'NA')
        )
    ) AS Ubicazione,
    j.UDC,
    j.Lotto,
    j.Prodotto,
    j.Descrizione
FROM JOINED j
WHERE 1=1
  AND ( :udc   IS NULL OR j.UDC            COLLATE Latin1_General_CI_AS LIKE CONCAT('%', :udc, '%') )
  AND ( :lotto IS NULL OR j.Lotto          COLLATE Latin1_General_CI_AS LIKE CONCAT('%', :lotto, '%') )
  AND ( :codice IS NULL OR j.Prodotto COLLATE Latin1_General_CI_AS LIKE CONCAT('%', :codice, '%') )
ORDER BY 
    CASE WHEN j.IDCella = 9999 THEN 1 ELSE 0 END,
    j.Corsia, j.Colonna, j.Fila, j.UDC, j.Lotto, j.Prodotto;
"""

class SearchWindow(tk.Toplevel):
    def __init__(self, parent: tk.Widget, db_app):
        super().__init__(parent)
        self.title("Warehouse · Ricerca UDC/Lotto/Codice")
        self.geometry("1100x720")
        self.minsize(900, 560)
        self.resizable(True, True)

        self.db = db_app
        self._busy = BusyOverlay(self)
        self._async = AsyncRunner(self)

        # stato ordinamento colonne (col -> reverse bool)
        self._sort_state: dict[str, bool] = {}

        self._build_ui()

    def _build_ui(self):
        # layout griglia principale
        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # --- barra di ricerca ---
        top = ttk.Frame(self)
        top.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)
        for i in range(8):
            top.grid_columnconfigure(i, weight=0)
        top.grid_columnconfigure(7, weight=1)

        ttk.Label(top, text="UDC:").grid(row=0, column=0, sticky="w")
        self.var_udc = tk.StringVar()
        e_udc = ttk.Entry(top, textvariable=self.var_udc, width=20)
        e_udc.grid(row=0, column=1, sticky="w", padx=(4, 12))

        ttk.Label(top, text="Lotto:").grid(row=0, column=2, sticky="w")
        self.var_lotto = tk.StringVar()
        e_lotto = ttk.Entry(top, textvariable=self.var_lotto, width=16)
        e_lotto.grid(row=0, column=3, sticky="w", padx=(4, 12))

        ttk.Label(top, text="Codice prodotto:").grid(row=0, column=4, sticky="w")
        self.var_codice = tk.StringVar()
        e_cod = ttk.Entry(top, textvariable=self.var_codice, width=20)
        e_cod.grid(row=0, column=5, sticky="w", padx=(4, 12))

        btn = ttk.Button(top, text="Cerca", command=self._do_search)
        btn.grid(row=0, column=6, sticky="w")

        btn_exp = ttk.Button(top, text="Esporta XLSX", command=self._export_xlsx)
        btn_exp.grid(row=0, column=7, sticky="e")

        # --- griglia risultati ---
        wrap = ttk.Frame(self)
        wrap.grid(row=1, column=0, sticky="nsew", padx=8, pady=(0, 8))
        wrap.grid_rowconfigure(0, weight=1)
        wrap.grid_columnconfigure(0, weight=1)

        # forza modalità Treeview (niente tksheet) per stabilità
        self.use_sheet = False

        cols = ("IDCella", "Ubicazione", "UDC", "Lotto", "Codice", "Descrizione")
        self.tree = ttk.Treeview(wrap, columns=cols, show="headings")
        # stile: zebra + header leggibile
        self._style = ttk.Style(self)
        try:
            self._style.theme_use(self._style.theme_use())
        except Exception:
            pass
        self._style.configure("Search.Treeview", rowheight=22, font=("", 9))
        self._style.configure("Search.Treeview.Heading", font=("", 9, "bold"), background="#F3F4F6")
        self._style.map("Search.Treeview", background=[("selected", "#DCEBFF")])
        self.tree.configure(style="Search.Treeview")
        # tag per righe alternate + id9999 evidenziate
        self.tree.tag_configure("even", background="#FFFFFF")
        self.tree.tag_configure("odd", background="#F7F9FC")
        # evidenzia spediti (IDCella=9999) in rosato tenue
        self.tree.tag_configure("id9999", background="#FFECEC", foreground="#B00020")

        sy = ttk.Scrollbar(wrap, orient="vertical", command=self.tree.yview)
        sx = ttk.Scrollbar(wrap, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        sy.grid(row=0, column=1, sticky="ns")
        sx.grid(row=1, column=0, sticky="ew")

        # doppio click → copia UDC; gestione click header
        self.tree.bind("<Double-1>", self._on_dclick)
        self.tree.bind("<Button-1>", self._maybe_handle_heading_click, add=True)
        self.tree.bind("<Double-1>", self._on_heading_double_click, add=True)

    def _apply_zebra(self):
        for i, iid in enumerate(self.tree.get_children("")):
            vals = self.tree.item(iid, "values")
            zebra = "even" if i % 2 == 0 else "odd"
            is9999 = False
            if vals:
                try:
                    is9999 = int(vals[0]) == 9999
                except Exception:
                    is9999 = False
            tags = ("id9999", zebra) if is9999 else (zebra,)
            self.tree.item(iid, tags=tags)

    # ---------------- AZIONI ----------------
    def _export_xlsx(self):
        # raccogli dati dalla griglia
        rows = []
        for iid in self.tree.get_children(""):
            rows.append(self.tree.item(iid, "values"))
        if not rows:
            messagebox.showinfo("Esporta", "Non ci sono righe da esportare.", parent=self)
            return
        if not _HAS_XLSX:
            messagebox.showerror("Esporta", "Per l'esportazione serve 'openpyxl' (pip install openpyxl).", parent=self)
            return
        # dialog salvataggio
        from datetime import datetime
        ts = datetime.now().strftime("%d_%m_%Y_%H-%M")
        default_name = f"esportazione_ricerca_{ts}.xlsx"
        fname = filedialog.asksaveasfilename(parent=self, title="Esporta in Excel",
                                             defaultextension=".xlsx",
                                             filetypes=[("Excel Workbook","*.xlsx")],
                                             initialfile=default_name)
        if not fname:
            return
        # crea workbook e scrivi
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Risultati"
            headers = ("IDCella","Ubicazione","UDC","Lotto","Codice","Descrizione")
            for j, h in enumerate(headers, start=1):
                c = ws.cell(row=1, column=j, value=h)
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
            r = 2
            for row in rows:
                for j, v in enumerate(row, start=1):
                    ws.cell(row=r, column=j, value=v)
                r += 1
            # autosize
            widths = {}
            for row in ws.iter_rows(values_only=True):
                for j, val in enumerate(row, start=1):
                    s = "" if val is None else str(val)
                    widths[j] = max(widths.get(j, 0), len(s))
            from openpyxl.utils import get_column_letter
            for j, w in widths.items():
                ws.column_dimensions[get_column_letter(j)].width = min(max(w + 2, 10), 60)
            wb.save(fname)
            messagebox.showinfo("Esporta", f"File creato:\n{fname}", parent=self)

        except Exception as ex:
            messagebox.showerror("Esporta", f"Errore durante l'esportazione:{ex}", parent=self)
    def _on_dclick(self, evt):
        # copia UDC solo se il doppio click avviene su una cella, non sull'header
        region = self.tree.identify("region", evt.x, evt.y)
        if region != "cell":
            return
        sel = self.tree.focus()
        if not sel:
            return
        vals = self.tree.item(sel, "values")
        if len(vals) >= 3 and vals[2]:  # UDC
            try:
                self.clipboard_clear()
                self.clipboard_append(vals[2])
            except Exception:
                pass

    # --------------- ORDINAMENTO COLONNE ---------------
    def _maybe_handle_heading_click(self, evt):
        # evita che il click sulle intestazioni selezioni una riga fantasma
        region = self.tree.identify("region", evt.x, evt.y)
        if region == "heading":
            return "break"

    def _on_heading_double_click(self, evt):
        # doppio click su intestazione: ordina la colonna corrispondente
        region = self.tree.identify("region", evt.x, evt.y)
        if region != "heading":
            return
        col_id = self.tree.identify_column(evt.x)  # es. '#1'
        try:
            idx = int(col_id.replace('#','')) - 1
        except Exception:
            return "break"
        cols = ("IDCella", "Ubicazione", "UDC", "Lotto", "Codice", "Descrizione")
        if 0 <= idx < len(cols):
            self._sort_by_column(cols[idx])
        return "break"

    def _sort_key_for_col(self, col: str, val: str):
        if val is None:
            return (1, "")  # None in fondo
        s = str(val)
        if col in ("IDCella",):
            # prova numero
            try:
                return (0, int(s))
            except Exception:
                return (0, s.lower())
        if col in ("Lotto", "Codice", "UDC"):
            return (0, s.lower())
        return (0, s.lower())

    def _sort_by_column(self, col: str):
        try:
            # toggle reverse
            rev = self._sort_state.get(col, False)
            self._sort_state[col] = not rev

            # raccogli dati correnti
            rows = []
            for iid in self.tree.get_children(""):
                vals = self.tree.item(iid, "values")
                row = {"iid": iid,
                       "IDCella": vals[0] if len(vals) > 0 else None,
                       "Ubicazione": vals[1] if len(vals) > 1 else None,
                       "UDC": vals[2] if len(vals) > 2 else None,
                       "Lotto": vals[3] if len(vals) > 3 else None,
                       "Codice": vals[4] if len(vals) > 4 else None,
                       "Descrizione": vals[5] if len(vals) > 5 else None}
                rows.append(row)

            rows.sort(key=lambda r: self._sort_key_for_col(col, r.get(col)), reverse=rev)

            def _apply_moves():
                for index, r in enumerate(rows):
                    self.tree.move(r["iid"], "", index)
                # aggiorna indicatori visuali nelle heading (▲/▼)
                for k in ("IDCella","Ubicazione","UDC","Lotto","Codice","Descrizione"):
                    base = {
                        "IDCella": "IDCella",
                        "Ubicazione": "Ubicazione",
                        "UDC": "UDC / Barcode",
                        "Lotto": "Lotto",
                        "Codice": "Codice prodotto",
                        "Descrizione": "Descrizione prodotto",
                    }[k]
                    if k == col:
                        arrow = " ▼" if not rev else " ▲"
                        self.tree.heading(k, text=base + arrow)
                    else:
                        self.tree.heading(k, text=base)
                self._apply_zebra()

            # posticipa le move per evitare re‑entrancy su doppio click
            self.after_idle(_apply_moves)
        except Exception:
            # in caso di problemi non lasciamo la finestra in stato incoerente
            pass

    # --- ordinamento per tksheet (doppio click sui titoli) ---
    def _on_sheet_header_double_click(self, event_dict):
        try:
            c = event_dict.get("column")
        except Exception:
            return
        if c is None:
            return
        headers = ["IDCella","Ubicazione","UDC","Lotto","Codice","Descrizione"]
        if not (0 <= c < len(headers)):
            return
        colname = headers[c]
        rev = self._sort_state.get(colname, False)
        self._sort_state[colname] = not rev

        data = self.sheet.get_sheet_data(return_copy=True)
        def keyf(row):
            val = row[c] if c < len(row) else None
            if val is None:
                return (1, "")
            s = str(val)
            if colname == "IDCella":
                try:
                    return (0, int(s))
                except Exception:
                    return (0, s.lower())
            return (0, s.lower())
        data.sort(key=keyf, reverse=rev)
        self.sheet.set_sheet_data(data)
        # evidenzia l'header ordinato (semplice: cambia testo temporaneamente)
        try:
            arrow = " ▼" if not rev else " ▲"
            hdrs = list(headers)
            hdrs[c] = hdrs[c] + arrow
            self.sheet.headers(hdrs)
        except Exception:
            pass

    def _do_search(self):
        udc = (self.var_udc.get() or "").strip()
        lotto = (self.var_lotto.get() or "").strip()
        codice = (self.var_codice.get() or "").strip()

        # Se nessun filtro, chiedi conferma (evita estrazione enorme non voluta)
        if not (udc or lotto or codice):
            if not messagebox.askyesno(
                "Conferma",
                "Nessun filtro impostato. Vuoi cercare su TUTTO il magazzino?",
                parent=self,
            ):
                return

        # Parametri: passa NULL se campo vuoto -> i filtri "si spengono"
        params = {
            "udc": (udc if udc else None),
            "lotto": (lotto if lotto else None),
            "codice": (codice if codice else None),
        }

        def _ok(res):
            rows = res.get("rows", []) if isinstance(res, dict) else []
            # --- popola UI ---
            if self.use_sheet:
                try:
                    data = []
                    for r in rows:
                        idc, ubi, udc_v, lot_v, cod_v, desc_v = r
                        data.append([idc, ubi, udc_v, lot_v, cod_v, desc_v])
                    self.sheet.set_sheet_data(data)
                    self.sheet.set_all_cell_sizes_to_text()
                except Exception as ex:
                    # fallback di sicurezza su Treeview
                    self.use_sheet = False
            if not self.use_sheet:
                # Treeview
                for iid in self.tree.get_children():
                    self.tree.delete(iid)
                for idx, r in enumerate(rows):
                    idc, ubi, udc_v, lot_v, cod_v, desc_v = r
                    zebra = "even" if idx % 2 == 0 else "odd"
                    try:
                        is9999 = int(idc) == 9999
                    except Exception:
                        is9999 = False
                    tags = ("id9999", zebra) if is9999 else (zebra,)
                    self.tree.insert("", "end", values=(idc, ubi, udc_v, lot_v, cod_v, desc_v), tags=tags)

            # --- feedback utente ---
            if not rows:
                messagebox.showinfo(
                    "Nessun risultato",
                    "Nessuna corrispondenza trovata con le chiavi di ricerca inserite.",
                    parent=self,
                )
            else:
                # reset campi se risultato non vuoto
                self.var_udc.set("")
                self.var_lotto.set("")
                self.var_codice.set("")
            self._busy.hide()

        def _err(ex):
            self._busy.hide()
            messagebox.showerror("Errore ricerca", str(ex), parent=self)

        self._async.run(self.db.query_json(SQL_SEARCH, params), _ok, _err, busy=self._busy, message="Cerco…")


def open_search_window(parent, db_app):
    key = "_search_window_singleton"
    ex = getattr(parent, key, None)
    if ex and ex.winfo_exists():
        try:
            ex.lift(); ex.focus_force(); return ex
        except Exception:
            pass
    w = SearchWindow(parent, db_app)
    setattr(parent, key, w)
    return w
