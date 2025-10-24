# layout_window.py
from __future__ import annotations
import tkinter as tk
from tkinter import ttk, Menu, messagebox, filedialog
from datetime import datetime

from gestione_aree_frame_async import BusyOverlay, AsyncRunner

# ---- Color palette ----
COLOR_EMPTY  = "#B0B0B0"  # grigio (vuota)
COLOR_FULL   = "#FFA500"  # arancione (una UDC)
COLOR_DOUBLE = "#D62728"  # rosso (>=2 UDC)
FG_DARK      = "#111111"
FG_LIGHT     = "#FFFFFF"


def pct_text(p_full: float, p_double: float | None = None) -> str:
    p_full = max(0.0, min(1.0, p_full))
    pf = round(p_full * 100, 1)
    pe = round(100 - pf, 1)
    if p_double and p_double > 0:
        pd = round(p_double * 100, 1)
        return f"Pieno {pf}%  ·  Vuoto {pe}%  (di cui doppie {pd}%)"
    return f"Pieno {pf}%  ·  Vuoto {pe}%"


class LayoutWindow(tk.Toplevel):
    """
    Visualizzazione layout corsie con matrice di celle.
    - Ogni cella è un pulsante colorato (vuota/piena/doppia)
    - Etichetta su DUE righe:
        1) "Corsia.Colonna.Fila" (una sola riga, senza andare a capo)
        2) barcode UDC (primo, se presente)
    - Ricerca per barcode UDC con cambio automatico corsia + highlight cella
    - Statistiche: globale e corsia selezionata
    - Export XLSX
    """
    def __init__(self, parent: tk.Widget, db_app):
        super().__init__(parent)
        self.title("Warehouse · Layout corsie")
        self.geometry("1200x740")
        self.minsize(980, 560)
        self.resizable(True, True)

        self.db = db_app
        self._busy  = BusyOverlay(self)
        self._async = AsyncRunner(self)

        # layout principale 5% / 80% / 15%
        self.grid_rowconfigure(0, weight=5)
        self.grid_rowconfigure(1, weight=80)
        self.grid_rowconfigure(2, weight=15)
        self.grid_columnconfigure(0, weight=1)

        # stato runtime
        self.corsia_selezionata = tk.StringVar()
        self.buttons: list[list[tk.Button]] = []
        self.btn_frames: list[list[tk.Frame]] = []
        self.state: list[list[int]] = []
        self.fila_txt: list[list[str]] = []
        self.col_txt: list[list[str]] = []
        self.desc: list[list[str]] = []
        self.udc1: list[list[str]] = []  # primo barcode UDC trovato (o "")

        # ricerca → focus differito (corsia, col, fila, barcode)
        self._pending_focus: tuple[str, str, str, str] | None = None
        self._highlighted: tuple[int, int] | None = None

        # anti-race: token per ignorare risposte vecchie
        self._req_counter = 0
        self._last_req = 0

        self._build_top()
        self._build_matrix_host()
        self._build_stats()

        self._load_corsie()
        self.bind("<Configure>", lambda e: self.after_idle(self._refresh_stats))

    # ---------------- TOP BAR ----------------
    def _build_top(self):
        top = ttk.Frame(self)
        top.grid(row=0, column=0, sticky="nsew", padx=8, pady=6)
        for i in range(4):
            top.grid_columnconfigure(i, weight=0)
        top.grid_columnconfigure(1, weight=1)

        # lista corsie
        lf = ttk.LabelFrame(top, text="Corsie")
        lf.grid(row=0, column=0, sticky="nsw")
        self.lb = tk.Listbox(lf, height=6, exportselection=False)
        self.lb.grid(row=0, column=0, sticky="nsw", padx=6, pady=4)
        self.lb.bind("<<ListboxSelect>>", self._on_select)

        # search by barcode
        srch = ttk.Frame(top)
        srch.grid(row=0, column=1, sticky="nsew", padx=(10, 10))
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(srch, textvariable=self.search_var, width=36)
        self.search_entry.grid(row=0, column=0, sticky="w")
        ttk.Button(srch, text="Cerca per barcode UDC", command=self._search_udc).grid(row=0, column=1, padx=(8, 0))
        srch.grid_columnconfigure(0, weight=1)

        # toolbar
        tb = ttk.Frame(top)
        tb.grid(row=0, column=3, sticky="ne")
        ttk.Button(tb, text="Aggiorna", command=self._refresh_current).grid(row=0, column=0, padx=4)
        ttk.Button(tb, text="Export XLSX", command=self._export_xlsx).grid(row=0, column=1, padx=4)

    # ---------------- MATRIX HOST ----------------
    def _build_matrix_host(self):
        center = ttk.Frame(self)
        center.grid(row=1, column=0, sticky="nsew", padx=8, pady=(0, 6))
        center.grid_rowconfigure(0, weight=1)
        center.grid_columnconfigure(0, weight=1)
        self.host = ttk.Frame(center, padding=4)
        self.host.grid(row=0, column=0, sticky="nsew")

    def _clear_highlight(self):
        if self._highlighted and self.buttons:
            r, c = self._highlighted
            try:
                if 0 <= r < len(self.buttons) and 0 <= c < len(self.buttons[r]):
                    btn = self.buttons[r][c]
                    if getattr(btn, "winfo_exists", None) and btn.winfo_exists():
                        try:
                            btn.configure(relief="raised", bd=1, highlightthickness=0)
                        except Exception:
                            pass
                    # clear blue frame border
                    try:
                        fr = self.btn_frames[r][c]
                        if fr and getattr(fr, "winfo_exists", None) and fr.winfo_exists():
                            fr.configure(highlightthickness=0)
                    except Exception:
                        pass
            except Exception:
                pass
        self._highlighted = None

    def _rebuild_matrix(self, rows: int, cols: int, state, fila_txt, col_txt, desc, udc1, corsia):
        # prima rimuovi highlight su vecchi bottoni
        self._clear_highlight()
        # ripulisci host
        for w in self.host.winfo_children():
            w.destroy()
        self.buttons.clear()
        self.btn_frames.clear()

        # salva matrici
        self.state, self.fila_txt, self.col_txt, self.desc, self.udc1 = state, fila_txt, col_txt, desc, udc1

        # ridistribuisci pesi griglia
        for r in range(rows):
            self.host.grid_rowconfigure(r, weight=1)
        for c in range(cols):
            self.host.grid_columnconfigure(c, weight=1)

        # crea Frame+Button per cella (righe invertite: fila "a" in basso)
        for r in range(rows):
            row_btns = []
            row_frames = []
            for c in range(cols):
                st = state[r][c]
                code = f"{corsia}.{col_txt[r][c]}.{fila_txt[r][c]}"  # PRIMA RIGA (in linea)
                udc  = udc1[r][c] or ""                              # SECONDA RIGA: barcode UDC
                text = f"{code}\n{udc}"
                cell = tk.Frame(self.host, bd=0, highlightthickness=0)
                btn = tk.Button(
                    cell, text=text, relief="raised", bd=1,
                    justify="center", wraplength=0  # wrap disattivato: niente 3a riga
                )
                if st == 0:
                    btn.configure(bg=COLOR_EMPTY,  fg=FG_DARK,
                                  activebackground="#9A9A9A", activeforeground=FG_DARK)
                elif st == 1:
                    btn.configure(bg=COLOR_FULL,   fg=FG_DARK,
                                  activebackground="#E69500", activeforeground=FG_DARK)
                else:
                    btn.configure(bg=COLOR_DOUBLE, fg=FG_LIGHT,
                                  activebackground="#B22222", activeforeground=FG_LIGHT)

                rr = (rows - 1) - r  # capovolgi
                cell.grid(row=rr, column=c, padx=1, pady=1, sticky="nsew")
                btn.pack(fill="both", expand=True)
                btn.configure(command=lambda rr=r, cc=c: self._open_menu(None, rr, cc))
                btn.bind("<Button-3>", lambda e, rr=r, cc=c: self._open_menu(e, rr, cc))
                row_btns.append(btn)
                row_frames.append(cell)
            self.buttons.append(row_btns)
            self.btn_frames.append(row_frames)

        # focus differito post-ricarica
        if self._pending_focus and self._pending_focus[0] == corsia:
            _, col, fila, _barcode = self._pending_focus
            self._pending_focus = None
            self._highlight_cell_by_labels(col, fila)

    # ---------------- CONTEXT MENU ----------------
    def _open_menu(self, event, r, c):
        st = self.state[r][c]
        corsia = self.corsia_selezionata.get()
        label = f"{corsia}.{self.col_txt[r][c]}.{self.fila_txt[r][c]}"
        m = Menu(self, tearoff=0)
        m.add_command(label="Apri dettaglio", command=lambda: self._toast(f"Dettaglio {label}"))
        if st == 0:
            m.add_command(label="Segna pieno",  command=lambda: self._set_cell(r, c, 1))
            m.add_command(label="Segna doppia", command=lambda: self._set_cell(r, c, 2))
        elif st == 1:
            m.add_command(label="Segna vuoto",  command=lambda: self._set_cell(r, c, 0))
            m.add_command(label="Segna doppia", command=lambda: self._set_cell(r, c, 2))
        else:
            m.add_command(label="Segna vuoto",  command=lambda: self._set_cell(r, c, 0))
            m.add_command(label="Segna pieno",  command=lambda: self._set_cell(r, c, 1))
        m.add_separator()
        m.add_command(label="Copia ubicazione", command=lambda: self._copy(label))
        x = self.winfo_pointerx() if event is None else event.x_root
        y = self.winfo_pointery() if event is None else event.y_root
        m.tk_popup(x, y)

    def _set_cell(self, r, c, val):
        self.state[r][c] = val
        btn = self.buttons[r][c]
        if val == 0:
            btn.configure(bg=COLOR_EMPTY,  fg=FG_DARK,  activebackground="#9A9A9A", activeforeground=FG_DARK)
        elif val == 1:
            btn.configure(bg=COLOR_FULL,   fg=FG_DARK,  activebackground="#E69500", activeforeground=FG_DARK)
        else:
            btn.configure(bg=COLOR_DOUBLE, fg=FG_LIGHT, activebackground="#B22222", activeforeground=FG_LIGHT)
        self._refresh_stats()

    # ---------------- STATS ----------------
    def _build_stats(self):
        bottom = ttk.Frame(self)
        bottom.grid(row=2, column=0, sticky="nsew", padx=8, pady=6)
        bottom.grid_columnconfigure(0, weight=1)

        ttk.Label(bottom, text="Riempimento globale", font=("", 10, "bold")).grid(row=0, column=0, sticky="w", pady=(0, 2))
        self.tot_canvas = tk.Canvas(bottom, height=22, highlightthickness=0)
        self.tot_canvas.grid(row=1, column=0, sticky="ew", padx=(0, 260))
        self.tot_text = ttk.Label(bottom, text=pct_text(0.0, 0.0))
        self.tot_text.grid(row=1, column=0, sticky="e")

        ttk.Label(bottom, text="Riempimento corsia selezionata", font=("", 10, "bold")).grid(row=2, column=0, sticky="w", pady=(10, 2))
        self.sel_canvas = tk.Canvas(bottom, height=22, highlightthickness=0)
        self.sel_canvas.grid(row=3, column=0, sticky="ew", padx=(0, 260))
        self.sel_text = ttk.Label(bottom, text=pct_text(0.0, 0.0))
        self.sel_text.grid(row=3, column=0, sticky="e")

        leg = ttk.Frame(bottom)
        leg.grid(row=4, column=0, sticky="w", pady=(10, 0))
        ttk.Label(leg, text="Legenda celle:").grid(row=0, column=0, padx=(0, 8))
        self._legend(leg, 1, "Vuota", COLOR_EMPTY)
        self._legend(leg, 3, "Piena", COLOR_FULL)
        self._legend(leg, 5, "Doppia UDC", COLOR_DOUBLE)

    def _legend(self, parent, col, text, color):
        box = tk.Canvas(parent, width=18, height=12, highlightthickness=0)
        box.create_rectangle(0, 0, 18, 12, fill=color, width=1, outline="#444")
        box.grid(row=0, column=col)
        ttk.Label(parent, text=text).grid(row=0, column=col + 1, padx=(4, 12))

    # ---------------- DATA LOADING ----------------
    def _load_corsie(self):
        sql = """
            WITH C AS (
                SELECT DISTINCT LTRIM(RTRIM(Corsia)) AS Corsia
                FROM dbo.Celle
                WHERE ID <> 9999 AND (DelDataOra IS NULL) AND LTRIM(RTRIM(Corsia)) <> '7G'
            )
            SELECT Corsia
            FROM C
            ORDER BY
              CASE
                WHEN LEFT(Corsia,3)='MAG' AND TRY_CONVERT(int, SUBSTRING(Corsia,4,50)) IS NOT NULL THEN 0
                WHEN TRY_CONVERT(int, Corsia) IS NOT NULL THEN 1
                ELSE 2
              END,
              CASE WHEN LEFT(Corsia,3)='MAG' THEN TRY_CONVERT(int, SUBSTRING(Corsia,4,50)) END,
              CASE WHEN TRY_CONVERT(int, Corsia) IS NOT NULL THEN TRY_CONVERT(int, Corsia) END,
              CASE WHEN TRY_CONVERT(int, Corsia) IS NOT NULL THEN SUBSTRING(Corsia, LEN(CAST(TRY_CONVERT(int, Corsia) AS varchar(20)))+1, 50) END,
              Corsia;
        """
        def _ok(res):
            rows = res.get("rows", []) if isinstance(res, dict) else []
            self.lb.delete(0, tk.END)
            corsie = [r[0] for r in rows]
            for c in corsie:
                self.lb.insert(tk.END, c)
            idx = corsie.index("1A") if "1A" in corsie else (0 if corsie else -1)
            if idx >= 0:
                self.lb.selection_clear(0, tk.END)
                self.lb.selection_set(idx)
                self.lb.see(idx)
                self._on_select(None)
            else:
                self._toast("Nessuna corsia trovata.")
            self._busy.hide()
        def _err(ex):
            self._busy.hide()
            messagebox.showerror("Errore", f"Caricamento corsie fallito:\n{ex}")
        self._async.run(self.db.query_json(sql, {}), _ok, _err, busy=self._busy, message="Carico corsie…")

    def _on_select(self, _):
        sel = self.lb.curselection()
        if not sel:
            return
        corsia = self.lb.get(sel[0])
        self.corsia_selezionata.set(corsia)
        self._load_matrix(corsia)

    def _select_corsia_in_listbox(self, corsia: str):
        for i in range(self.lb.size()):
            if self.lb.get(i) == corsia:
                self.lb.selection_clear(0, tk.END)
                self.lb.selection_set(i)
                self.lb.see(i)
                break

    def _load_matrix(self, corsia: str):
        # nuovo token richiesta → evita che risposte vecchie spazzino la UI
        self._req_counter += 1
        req_id = self._req_counter
        self._last_req = req_id

        sql = """
        WITH C AS (
            SELECT
                ID,
                LTRIM(RTRIM(Corsia))  AS Corsia,
                LTRIM(RTRIM(Fila))    AS Fila,
                LTRIM(RTRIM(Colonna)) AS Colonna,
                Descrizione
            FROM dbo.Celle
            WHERE ID <> 9999 AND (DelDataOra IS NULL)
              AND LTRIM(RTRIM(Corsia)) <> '7G' AND LTRIM(RTRIM(Corsia)) = :corsia
        ),
        R AS (
            SELECT Fila,
                   DENSE_RANK() OVER (
                     ORDER BY CASE WHEN TRY_CONVERT(int, Fila) IS NULL THEN 1 ELSE 0 END,
                              TRY_CONVERT(int, Fila), Fila
                   ) AS RowN
            FROM C GROUP BY Fila
        ),
        K AS (
            SELECT Colonna,
                   DENSE_RANK() OVER (
                     ORDER BY CASE WHEN TRY_CONVERT(int, Colonna) IS NULL THEN 1 ELSE 0 END,
                              TRY_CONVERT(int, Colonna), Colonna
                   ) AS ColN
            FROM C GROUP BY Colonna
        ),
        S AS (
            SELECT c.ID, COUNT(DISTINCT g.BarcodePallet) AS n
            FROM C AS c
            LEFT JOIN dbo.XMag_GiacenzaPallet AS g ON g.IDCella = c.ID
            GROUP BY c.ID
        ),
        U AS (
            SELECT c.ID, MIN(g.BarcodePallet) AS FirstUDC
            FROM C c
            LEFT JOIN dbo.XMag_GiacenzaPallet g ON g.IDCella = c.ID
            GROUP BY c.ID
        )
        SELECT
            r.RowN, k.ColN,
            CASE WHEN s.n IS NULL OR s.n = 0 THEN 0
                 WHEN s.n = 1 THEN 1
                 ELSE 2 END AS Stato,
            c.Descrizione,
            LTRIM(RTRIM(c.Fila)) AS FilaTxt,
            LTRIM(RTRIM(c.Colonna)) AS ColTxt,
            U.FirstUDC
        FROM C c
        JOIN R r ON r.Fila = c.Fila
        JOIN K k ON k.Colonna = c.Colonna
        LEFT JOIN S s ON s.ID = c.ID
        LEFT JOIN U ON U.ID = c.ID
        ORDER BY r.RowN, k.ColN;
        """
        def _ok(res):
            # ignora risposte superate
            if req_id < self._last_req:
                return
            rows = res.get("rows", []) if isinstance(res, dict) else []
            if not rows:
                # mostra matrice vuota senza rimuovere il frame (evita "schermo bianco")
                self._rebuild_matrix(0, 0, [], [], [], [], [], corsia)
                self._refresh_stats()
                self._busy.hide()
                return
            max_r = max_c = 0
            for row in rows:
                rown, coln = row[0], row[1]
                if rown and coln:
                    max_r = max(max_r, int(rown))
                    max_c = max(max_c, int(coln))
            mat  = [[0] * max_c for _ in range(max_r)]
            fila = [[""] * max_c for _ in range(max_r)]
            col  = [[""] * max_c for _ in range(max_r)]
            desc = [[""] * max_c for _ in range(max_r)]
            udc  = [[""] * max_c for _ in range(max_r)]
            for row in rows:
                rown, coln, stato, descr, fila_txt, col_txt, first_udc = row
                r = int(rown) - 1
                c = int(coln) - 1
                mat[r][c]  = int(stato)
                fila[r][c] = str(fila_txt or "")
                col[r][c]  = str(col_txt or "")
                desc[r][c] = str(descr or f"{corsia}.{col_txt}.{fila_txt}")
                udc[r][c]  = str(first_udc or "")
            self._rebuild_matrix(max_r, max_c, mat, fila, col, desc, udc, corsia)
            self._refresh_stats()
            self._busy.hide()
        def _err(ex):
            if req_id < self._last_req:
                return
            self._busy.hide()
            messagebox.showerror("Errore", f"Caricamento matrice {corsia} fallito:\n{ex}")
        self._async.run(self.db.query_json(sql, {"corsia": corsia}), _ok, _err, busy=self._busy, message=f"Carico corsia {corsia}…")

    # ---------------- SEARCH ----------------
    def _search_udc(self):
        barcode = (self.search_var.get() or "").strip()
        if not barcode:
            self._toast("Inserisci un barcode UDC da cercare.")
            return

        # bump token per impedire che una vecchia _load_matrix cancelli UI
        self._req_counter += 1
        search_req_id = self._req_counter
        self._last_req = search_req_id

        sql = """
            SELECT TOP (1)
                   RTRIM(c.Corsia)  AS Corsia,
                   RTRIM(c.Colonna) AS Colonna,
                   RTRIM(c.Fila)    AS Fila,
                   c.ID             AS IDCella
            FROM dbo.XMag_GiacenzaPallet g
            JOIN dbo.Celle c ON c.ID = g.IDCella
            WHERE g.BarcodePallet = :barcode
              AND c.ID <> 9999 AND RTRIM(c.Corsia) <> '7G'
        """
        def _ok(res):
            if search_req_id < self._last_req:
                return
            rows = res.get("rows", []) if isinstance(res, dict) else []
            if not rows:
                messagebox.showinfo("Ricerca", f"UDC {barcode} non trovata.", parent=self)
                return
            corsia, col, fila, _idc = rows[0]
            corsia = str(corsia).strip(); col = str(col).strip(); fila = str(fila).strip()
            self._pending_focus = (corsia, col, fila, barcode)

            # sincronizza listbox e carica SEMPRE la corsia della UDC
            self._select_corsia_in_listbox(corsia)
            self.corsia_selezionata.set(corsia)
            self._load_matrix(corsia)  # highlight avverrà in _rebuild_matrix
        def _err(ex):
            if search_req_id < self._last_req:
                return
            messagebox.showerror("Ricerca", f"Errore ricerca UDC:\n{ex}", parent=self)

        self._async.run(self.db.query_json(sql, {"barcode": barcode}), _ok, _err, busy=self._busy, message="Cerco UDC…")

    def _try_highlight(self, col_txt: str, fila_txt: str) -> bool:
        for r in range(len(self.col_txt)):
            for c in range(len(self.col_txt[r])):
                if self.col_txt[r][c] == col_txt and self.fila_txt[r][c] == fila_txt:
                    self._clear_highlight()
                    btn = self.buttons[r][c]
                    btn.configure(relief="sunken", bd=3)
                    try:
                        fr = self.btn_frames[r][c]
                        fr.configure(highlightbackground="blue", highlightcolor="blue", highlightthickness=2)
                    except Exception:
                        pass
                    self._highlighted = (r, c)
                    return True
        return False

    def _highlight_cell_by_labels(self, col_txt: str, fila_txt: str):
        if not self._try_highlight(col_txt, fila_txt):
            self._toast("Cella trovata ma non mappabile a pulsante.")

    # ---------------- COMMANDS ----------------
    def _refresh_current(self):
        if self.corsia_selezionata.get():
            self._load_matrix(self.corsia_selezionata.get())

    def _export_xlsx(self):
        if not self.state:
            messagebox.showinfo("Export", "Nessuna matrice da esportare.")
            return
        corsia = self.corsia_selezionata.get() or "NA"
        ts = datetime.now().strftime("%d_%m_%Y_%H-%M")
        default = f"layout_matrice_{corsia}_{ts}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Esporta matrice",
            defaultextension=".xlsx",
            initialfile=default,
            filetypes=[("Excel", "*.xlsx")]
        )
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Alignment, Font
        except Exception as ex:
            messagebox.showerror("Export", f"Manca openpyxl: {ex}\nInstalla con: pip install openpyxl")
            return
        rows = len(self.state)
        cols = len(self.state[0]) if self.state else 0
        wb = Workbook()
        ws1 = wb.active
        ws1.title = f"Dettaglio {corsia}"
        ws1.append(["Corsia", "FilaIdx", "ColIdx", "Stato", "Descrizione", "FilaTxt", "ColTxt", "UDC1"])
        for r in range(rows):
            for c in range(cols):
                st = self.state[r][c]
                stato_lbl = "Vuota" if st == 0 else ("Piena" if st == 1 else "Doppia")
                ws1.append([corsia, r + 1, c + 1, stato_lbl,
                            self.desc[r][c], self.fila_txt[r][c], self.col_txt[r][c], self.udc1[r][c]])
        for cell in ws1[1]:
            cell.font = Font(bold=True)

        ws2 = wb.create_sheet(f"Matrice {corsia}")
        fills = {
            0: PatternFill("solid", fgColor="B0B0B0"),
            1: PatternFill("solid", fgColor="FFA500"),
            2: PatternFill("solid", fgColor="D62728"),
        }
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r in range(rows):
            for c in range(cols):
                value = f"{corsia}.{self.col_txt[r][c]}.{self.fila_txt[r][c]}\n{self.udc1[r][c]}"
                cell = ws2.cell(row=(rows - r), column=c + 1, value=value)  # capovolto per avere 'a' in basso
                cell.fill = fills.get(self.state[r][c], fills[0])
                cell.alignment = center
        try:
            wb.save(path)
            self._toast(f"Esportato: {path}")
        except Exception as ex:
            messagebox.showerror("Export", f"Salvataggio fallito:\n{ex}")

    # ---------------- STATS ----------------
    def _refresh_stats(self):
        # globale dal DB
        sql_tot = """
            WITH C AS (
                SELECT ID
                FROM dbo.Celle
                WHERE ID <> 9999 AND (DelDataOra IS NULL)
                  AND LTRIM(RTRIM(Corsia)) <> '7G'
                  AND LTRIM(RTRIM(Fila)) IS NOT NULL
                  AND LTRIM(RTRIM(Colonna)) IS NOT NULL
            ),
            S AS (
                SELECT c.ID, COUNT(DISTINCT g.BarcodePallet) AS n
                FROM C AS c LEFT JOIN dbo.XMag_GiacenzaPallet AS g ON g.IDCella = c.ID
                GROUP BY c.ID
            )
            SELECT
              CAST(SUM(CASE WHEN s.n>0 THEN 1 ELSE 0 END) AS float)/NULLIF(COUNT(*),0) AS PercPieno,
              CAST(SUM(CASE WHEN s.n>1 THEN 1 ELSE 0 END) AS float)/NULLIF(COUNT(*),0) AS PercDoppie
            FROM C LEFT JOIN S s ON s.ID = C.ID;
        """
        def _ok(res):
            rows = res.get("rows", []) if isinstance(res, dict) else []
            p_full = float(rows[0][0] or 0.0) if rows else 0.0
            p_dbl  = float(rows[0][1] or 0.0) if rows else 0.0
            self._draw_bar(self.tot_canvas, p_full)
            self.tot_text.configure(text=pct_text(p_full, p_dbl))
        self._async.run(self.db.query_json(sql_tot, {}), _ok, lambda e: None, busy=None, message=None)

        # selezionata dalla matrice in memoria
        if self.state:
            tot = sum(len(r) for r in self.state)
            full = sum(1 for row in self.state for v in row if v in (1, 2))
            doubles = sum(1 for row in self.state for v in row if v == 2)
            p_full = (full / tot) if tot else 0.0
            p_dbl  = (doubles / tot) if tot else 0.0
        else:
            p_full = p_dbl = 0.0
        self._draw_bar(self.sel_canvas, p_full)
        self.sel_text.configure(text=pct_text(p_full, p_dbl))

    def _draw_bar(self, cv: tk.Canvas, p_full: float):
        cv.delete("all")
        w = max(300, cv.winfo_width() or 600)
        h = 18
        fw = int(w * max(0.0, min(1.0, p_full)))
        cv.create_rectangle(2, 2, 2 + fw, 2 + h, fill="#D62728", width=0)
        cv.create_rectangle(2 + fw, 2, 2 + w, 2 + h, fill="#2CA02C", width=0)
        cv.create_rectangle(2, 2, 2 + w, 2 + h, outline="#555", width=1)

    # ---------------- UTIL ----------------
    def _toast(self, msg, ms=1400):
        if not hasattr(self, "_status"):
            self._status = ttk.Label(self, relief="groove", anchor="w")
            self._status.grid(row=3, column=0, sticky="ew")
        self._status.configure(text=msg)
        self.after(ms, lambda: self._status.configure(text=""))

    def _copy(self, txt: str):
        self.clipboard_clear()
        self.clipboard_append(txt)
        self._toast(f"Copiato: {txt}")


def open_layout_window(parent, db_app):
    key = "_layout_window_singleton"
    ex = getattr(parent, key, None)
    if ex and ex.winfo_exists():
        try:
            ex.lift()
            ex.focus_force()
            return ex
        except Exception:
            pass
    w = LayoutWindow(parent, db_app)
    setattr(parent, key, w)
    return w
